import os
import sys
import sqlite3
import optparse
import hashlib
import socket

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# NOTE: This code is not really mine. It is taken from the solution from the manning liveproject

def run(db_file, folder):
    '''Run the program'''
    # For now just a dummy
    setup_tracking_table("afile.txt", "1234", db_file)


def setup_tracking_table(fname, md5, db_file):
    '''Setup the files table'''
    create_tracking_table(db_file)
    create_tracking_table_idx(db_file)
    insert_tracking_table(fname, md5, db_file)


def create_db(db_file):
    '''Create the database if it doesn't exist'''
    try:
        conn = sqlite3.connect(db_file, timeout=2)
    except BaseException as err:
        print(str(err))
        conn = None
    return conn


def md5_in_db(db_file, fname):
    items = []
    query = "select md5 from files where fname = ?"
    try:
        conn = create_db(db_file)
        if conn:
            if table_exists("files", conn):
                try:
                    cur = conn.cursor()
                    cur.execute(query, (fname,))
                    for row in cur:
                        items.append(row[0])
                except sqlite3.OperationalError as err:
                    if cur:
                        cur.close()
                finally:
                    if cur:
                        cur.close()
    except sqlite3.OperationalError as err:
        if conn:
            conn.close()
    finally:
        if conn:
            conn.close()
    return items


def create_tracking_table(db_file):
    '''Create the table'''
    result = False
    query = "create table files(file text, md5 text)"
    try:
        conn = create_db(db_file)
        if conn is not None:
            if not table_exists("files", conn):
                try:
                    cur = conn.cursor()
                    cur.execute(query)
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cur:
                        cur.close()
                finally:
                    conn.commit()
                    if cur:
                        cur.close()
                    result = True
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn:
            conn.close()
    finally:
        if conn:
            conn.close()
    return result


def create_tracking_table_idx(db_file):
    '''Create an indec for the table'''
    table = "files"
    query = "create index idx on files (file, md5)"
    run_query(db_file, query, None)


def run_query(db_file, query, args):
    '''Run the query'''
    conn = create_db(db_file)

    try:
        if conn is not None:
            if table_exists("files", conn):
                cur = conn.cursor()
                try:
                    if args:
                        cur.execute(query, args)
                    else:
                        cur.execute(query)
                except sqlite3.OperationalError as err:
                    print(str(err))
                    if cur:
                        cur.close()
                finally:
                    conn.commit()
                    if cur:
                        cur.close()
                    result = True
    except sqlite3.OperationalError as err:
        print(str(err))
        if conn:
            conn.close()
    finally:
        if conn:
            conn.close()


def update_tracking_table(fname, md5, db_file):
    '''Update the files table'''

    query = "update files set md5=? where file = ?"
    args = (md5, fname)
    run_query(db_file, query, args)


def insert_tracking_table(fname, md5, db_file):
    '''Insert into the files table'''
    query = "insert into files(file, md5) values(?,?)"
    args = (fname, md5)
    run_query(db_file, query, args)


def cursor(conn, query, args):
    '''Create a cursor and run the query'''
    result = False
    cur = conn.cursor()

    try:
        cur.execute(query, args)
        rows = cur.fetchall()
        if len(rows) > 0:
            result = True
    except sqlite3.OperationalError as err:
        print(str(err))
        if cur != None:
            cur.close()
    return result


def table_exists(table, conn):
    '''Check if a table exists'''
    result = False
    query = "select name from sqlite_master where type = 'table' and name = ?"
    args = (table,)
    result = cursor(conn, query, args)
    return result


def has_changed(fname, md5, db_file):
    '''Check if a file has changed'''
    md5s = md5_in_db(db_file, fname)
    if len(md5s) == 1:
        if md5s[0] != md5:
            update_tracking_table(fname, md5, db_file)
        else:
            setup_tracking_table(fname, md5, db_file)


def get_fileext(fname):
    '''Get the filename extension'''
    ext = os.path.splitext(fname)[1]
    return ext


def get_mod_date(fname):
    '''Get the file modified date'''
    try:
        mtime = os.path.getmtime(fname)
    except:
        mtime = 0
    return mtime


def md5_short(fname):
    '''Get the md5 file hash'''
    md5_hash = hashlib.md5()
    with open(fname, 'r') as fin:
        while True:
            buff = fin.buffer(8192)
            if not buff:
                break
            md5_hash.update(buff)
        md5_digest = md5_hash.hexdigest()
    md5 = str(md5_digest).lower()
    return md5


def load_fields():
    flds = []
    ext = []
    config = "./filechanges.ini"
    if os.path.isfile(config):
        cfile = open(config, 'r')
        # Parse each config file line and get the folder and extensions
        for line in cfile:
            folder, exts = line.split("|")
            if len(exts) > 0:
                for ex in exts.split(","):
                    ext.append(ex)
            if len(folder) > 0:
                flds.append(folder)

    return flds, ext


def checkfilechanges(folder, exclude, ws):
    changed = False
    """Checks for files changes"""
    for subdir, dirs, files in os.walk(folder):
        for fname in files:
            origin = os.path.join(subdir, fname)
            if os.path.isfile(origin):
                fext = get_fileext(origin)
                if not fext in exclude:
                    md5 = md5_short(origin)
                    if has_changed(origin, md5):
                        changed = True
    return changed


def runfilechanges(ws):
    changed = False
    fldexts = load_fields()
    for i, fld in enumerate(fldexts[0]):
        exts = fldexts[1]
        changed = checkfilechanges(fld, exts[i], ws)
    return changed


def startxlsreport():
    wb = Workbook()
    ws = wb.active
    ws.title = socket.gethostname()
    st = getdt("%d-%b-%Y %H_%M_%S")
    return wb, ws, st

def getdt(fmt):
    today = datetime.today()
    return today.strftime(fmt)

def endxlsreport(wb, st):
    dt = ' from ' + st + ' to ' + getdt("%d-%b-%Y %H_%M_%S")
    fn = "testfile" + dt + '.xlsx'
    wb.save(fn)

def headerxlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File Name")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    a1 = ws['A1']
    b1 = ws['B1']
    c1 = ws['C1']
    d1 = ws['D1']
    e1 = ws['E1']

    a1.font = Font(color="000000", bold=True)
    b1.font = Font(color="000000", bold=True)
    c1.font = Font(color="000000", bold=True)
    d1.font = Font(color="000000", bold=True)
    e1.font = Font(color="000000", bold=True)

def getlastrow(ws):
    rw = 1
    for cell in ws["A"]:
        if cell.value is None:
            break
        else:
            rw += 1
    return rw

def rowxlsreport(ws, fn, ffn, fld, d, t):
    rw = getlastrow(ws)

    ws.cell(row=rw, column=1, value=fn)
    ws.cell(row=rw, column=2, value=ffn)
    ws.cell(row=rw, column=3, value=fld)
    ws.cell(row=rw, column=4, value=d)
    ws.cell(row=rw, column=5, value=t)


def execute(args):
    chn = 0
    if len(args) > 1:
        if args[1].lower() == '--loop':
            # To be done in the last milestone
            wb, ws, st = startxlsreport()
            try:
                while True:
                    changed = runfilechanges(ws)
                    if changed:
                        chn += 1
            except KeyboardInterrupt:
                print('Stopped...')
                if chn > 0:
                    # To be done in the last milestone
                    endxlsreport(wb, st)
    else:
        # To be done in the last milestone
        wb, ws, st = startxlsreport()
        changed = runfilechanges(ws)
        if changed:
            # To be done in the last milestone
            endxlsreport(wb, st)

if __name__ == '__main__':
    execute(sys.argv)
